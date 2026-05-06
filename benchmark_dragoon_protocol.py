"""
Protocol-level benchmark for Dragoon Lv-1 / Lv-2 / Lv-3 (per HeteroRA Dragoon).

Compared to ``benchmark_dragoon.py`` (sub-algorithm breakdown), this script reports
composed phases aligned with evaluation tables:

- **Setup**: long-term keys installed before an online attestation (AttKGen + VerKGen;
  VerKGen is SIG.KGen for Lv-1; IBS.Setup for Lv-2/Lv-3).
- **Attest**: attester-side evidence generation (Lv-1/Lv-2: SIG.Sign on ``m``; Lv-3:
  BlGen + BlPubKey + BlSign per Fig. 3).
- **Verify**: completion of the attestation result on verifier/proxy side:
  Lv-1: VerSign + FinVf (no delegation); Lv-2/Lv-3: ProxKGen + ProxSign + FinVf
  (delegated RA service path).

Also records byte sizes: long-term attester ``pk``, evidence identity ``bpk`` (equals ``pka``
when no blinding), proxy delegation key ``rk`` (IBS user secret; Lv-1 uses 0), and the two
signatures σ_attester (KBS/DL signature on the report) and σ_final (verifier SIG on Lv-1,
IBS signature on Lv-2/Lv-3).

Outputs **three** workbooks (one per construction level):

- ``dragoon_protocol_lv1_benchmark.xlsx``
- ``dragoon_protocol_lv2_benchmark.xlsx``
- ``dragoon_protocol_lv3_benchmark.xlsx``

Each workbook uses multi-sheet rows by security profile (P-224 … P-521). Plus copies where
configured.
"""
from __future__ import annotations

import shutil
import statistics
import time
from pathlib import Path
from typing import Callable, List

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from dragoon_lv1 import DragoonLv1, DragoonLv1Evidence
from dragoon_lv2 import DragoonLv2, DragoonLv2Evidence
from dragoon_lv3 import (
    DragoonLv3,
    DragoonLv3Evidence,
    IBSDLUserSk,
    _point_bytes,
    serialize_ibs_sig,
    serialize_kbs_sig,
)


def _mean_ms(samples: List[float]) -> float:
    return statistics.mean(samples) * 1000.0


def _stdev_ms(samples: List[float]) -> float:
    if len(samples) < 2:
        return 0.0
    return statistics.stdev(samples) * 1000.0


def _bench_op(op: Callable[[], None], iterations: int) -> tuple[float, float]:
    samples: List[float] = []
    for _ in range(iterations):
        t0 = time.perf_counter()
        op()
        samples.append(time.perf_counter() - t0)
    return _mean_ms(samples), _stdev_ms(samples)


def _ibs_user_sk_bytes(usk: IBSDLUserSk, order: int) -> int:
    """Canonical byte length for proxy key material (identity + Extract outputs)."""
    bl = (order.bit_length() + 7) // 8
    return len(usk.id_bytes) + len(_point_bytes(usk.R)) + bl


def _security_sheets() -> dict[str, tuple[str, dict]]:
    return {
        "lambda~112_P-224": ("SECP224R1", {"symmetric_security_bits_approx": 112}),
        "lambda~128_P-256": ("SECP256R1", {"symmetric_security_bits_approx": 128}),
        "lambda~192_P-384": ("SECP384R1", {"symmetric_security_bits_approx": 192}),
        "lambda~256_P-521": ("SECP521R1", {"symmetric_security_bits_approx": 256}),
    }


def bench_lv1_curve(curve_name: str, iterations: int, warmup: int) -> tuple[dict, dict]:
    m = b"tee-report-hash-placeholder-32bytes!!"
    d1 = DragoonLv1.create(curve_name)

    def setup_lv1():
        d1.att_kgen()
        d1.ver_kgen()

    for _ in range(warmup):
        setup_lv1()
    s_mean, s_sd = _bench_op(setup_lv1, iterations)

    att_keys = d1.att_kgen()
    ev1 = d1.attest(att_keys.sk, att_keys.pk, m)

    def attest_lv1():
        d1.attest(att_keys.sk, att_keys.pk, m)

    for _ in range(warmup):
        attest_lv1()
    a_mean, a_sd = _bench_op(attest_lv1, iterations)

    ver1 = d1.ver_kgen()
    ev1_fix = DragoonLv1Evidence(pka=ev1.pka, m=m, sigma_a=ev1.sigma_a)
    sf1 = d1.ver_sign(ver1.sk, ev1_fix)
    assert sf1 is not None

    def verify_lv1():
        sff = d1.ver_sign(ver1.sk, ev1_fix)
        assert sff is not None
        d1.fin_vf(ver1.pk, ev1_fix, sff)

    for _ in range(warmup):
        verify_lv1()
    v_mean, v_sd = _bench_op(verify_lv1, iterations)

    pk_b = len(_point_bytes(att_keys.pk))
    row = {
        "level": "Lv-1",
        "setup_mean_ms": s_mean,
        "setup_stdev_ms": s_sd,
        "attest_mean_ms": a_mean,
        "attest_stdev_ms": a_sd,
        "verify_mean_ms": v_mean,
        "verify_stdev_ms": v_sd,
        "pk_bytes": pk_b,
        "bpk_bytes": pk_b,
        "rk_bytes": 0,
        "sigma_attester_bytes": len(serialize_kbs_sig(d1.kbs.curve, ev1.sigma_a)),
        "sigma_final_bytes": len(serialize_kbs_sig(d1.kbs.curve, sf1)),
        "bpk_note": "bpk = pka (vanilla RA)",
        "rk_note": "ProxKGen unused",
    }
    meta = {
        "curve": curve_name,
        "order_bits": d1.kbs.n.bit_length(),
        "hash_kbs": d1.kbs._hash,
        "iterations": iterations,
        "warmup": warmup,
    }
    return row, meta


def bench_lv2_curve(curve_name: str, iterations: int, warmup: int) -> tuple[dict, dict]:
    m = b"tee-report-hash-placeholder-32bytes!!"
    d2 = DragoonLv2.create(curve_name)

    def setup_lv2():
        d2.att_kgen()
        d2.ver_kgen()

    for _ in range(warmup):
        setup_lv2()
    s_mean, s_sd = _bench_op(setup_lv2, iterations)

    att2 = d2.att_kgen()
    ev2 = d2.attest(att2.sk, att2.pk, m)

    def attest_lv2():
        d2.attest(att2.sk, att2.pk, m)

    for _ in range(warmup):
        attest_lv2()
    a_mean, a_sd = _bench_op(attest_lv2, iterations)

    msk2 = d2.ver_kgen()
    ev2_fix = DragoonLv2Evidence(pka=ev2.pka, m=m, sigma_a=ev2.sigma_a)
    rk2 = d2.prox_kgen(msk2, ev2_fix.pka)
    sf2 = d2.prox_sign(rk2, ev2_fix, msk2)
    assert sf2 is not None

    def verify_lv2():
        rk = d2.prox_kgen(msk2, ev2_fix.pka)
        sff = d2.prox_sign(rk, ev2_fix, msk2)
        assert sff is not None
        d2.fin_vf(msk2, ev2_fix, sff)

    for _ in range(warmup):
        verify_lv2()
    v_mean, v_sd = _bench_op(verify_lv2, iterations)

    row = {
        "level": "Lv-2",
        "setup_mean_ms": s_mean,
        "setup_stdev_ms": s_sd,
        "attest_mean_ms": a_mean,
        "attest_stdev_ms": a_sd,
        "verify_mean_ms": v_mean,
        "verify_stdev_ms": v_sd,
        "pk_bytes": len(_point_bytes(att2.pk)),
        "bpk_bytes": len(_point_bytes(ev2_fix.pka)),
        "rk_bytes": _ibs_user_sk_bytes(rk2, d2.kbs.n),
        "sigma_attester_bytes": len(serialize_kbs_sig(d2.kbs.curve, ev2.sigma_a)),
        "sigma_final_bytes": len(serialize_ibs_sig(d2.kbs.curve, sf2)),
        "bpk_note": "bpk = pka (delegated RA, Fig.2)",
        "rk_note": "IBS.Extract(msk, id=pka)",
    }
    meta = {
        "curve": curve_name,
        "order_bits": d2.kbs.n.bit_length(),
        "hash_kbs": d2.kbs._hash,
        "iterations": iterations,
        "warmup": warmup,
    }
    return row, meta


def bench_lv3_curve(curve_name: str, iterations: int, warmup: int) -> tuple[dict, dict]:
    m = b"tee-report-hash-placeholder-32bytes!!"
    d3 = DragoonLv3.create(curve_name)

    def setup_lv3():
        d3.att_kgen()
        d3.ver_kgen()

    for _ in range(warmup):
        setup_lv3()
    s_mean, s_sd = _bench_op(setup_lv3, iterations)

    att3 = d3.att_kgen()

    def attest_lv3():
        d3.attest(att3.sk, att3.pk, m)

    for _ in range(warmup):
        attest_lv3()
    a_mean, a_sd = _bench_op(attest_lv3, iterations)

    bpk3, sig_a3, _bk = d3.attest(att3.sk, att3.pk, m)
    ev3_fix = DragoonLv3Evidence(bpk=bpk3, m=m, sigma_a=sig_a3)
    msk3 = d3.ver_kgen()
    rk3 = d3.prox_kgen(msk3, ev3_fix.bpk)
    sf3 = d3.prox_sign(rk3, ev3_fix, msk3)
    assert sf3 is not None

    def verify_lv3():
        rk = d3.prox_kgen(msk3, ev3_fix.bpk)
        sff = d3.prox_sign(rk, ev3_fix, msk3)
        assert sff is not None
        d3.fin_vf(msk3, ev3_fix, sff)

    for _ in range(warmup):
        verify_lv3()
    v_mean, v_sd = _bench_op(verify_lv3, iterations)

    row = {
        "level": "Lv-3",
        "setup_mean_ms": s_mean,
        "setup_stdev_ms": s_sd,
        "attest_mean_ms": a_mean,
        "attest_stdev_ms": a_sd,
        "verify_mean_ms": v_mean,
        "verify_stdev_ms": v_sd,
        "pk_bytes": len(_point_bytes(att3.pk)),
        "bpk_bytes": len(_point_bytes(bpk3)),
        "rk_bytes": _ibs_user_sk_bytes(rk3, d3.kbs.n),
        "sigma_attester_bytes": len(serialize_kbs_sig(d3.kbs.curve, sig_a3)),
        "sigma_final_bytes": len(serialize_ibs_sig(d3.kbs.curve, sf3)),
        "bpk_note": "bpk = BlPubKey(pka,bk)",
        "rk_note": "IBS.Extract(msk, id=bpk)",
    }
    meta = {
        "curve": curve_name,
        "order_bits": d3.kbs.n.bit_length(),
        "hash_kbs": d3.kbs._hash,
        "iterations": iterations,
        "warmup": warmup,
    }
    return row, meta


def write_protocol_workbook(
    path: Path,
    title_line: str,
    protocol_notes: tuple[str, str, str],
    all_results: dict[str, tuple[list[dict], dict]],
) -> None:
    wb = Workbook()
    first = True
    for sheet_name, (rows, meta) in all_results.items():
        if first:
            ws = wb.active
            ws.title = sheet_name[:31]
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name[:31])

        ws["A1"] = title_line
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
        ws["A1"].font = Font(bold=True)

        r = 3
        ws.cell(r, 1, "Parameter")
        ws.cell(r, 2, "Value")
        r += 1
        for k, v in meta.items():
            ws.cell(r, 1, str(k))
            ws.cell(r, 2, v if not isinstance(v, float) else round(v, 6))
            r += 1

        r += 1
        for note in protocol_notes:
            ws.cell(r, 1, note)
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
            r += 1
        r += 1

        headers = [
            "level",
            "setup_mean_ms",
            "setup_stdev_ms",
            "attest_mean_ms",
            "attest_stdev_ms",
            "verify_mean_ms",
            "verify_stdev_ms",
            "pk_bytes",
            "bpk_bytes",
            "rk_bytes",
            "sigma_attester_bytes",
            "sigma_final_bytes",
            "bpk_note",
            "rk_note",
        ]
        for c, h in enumerate(headers, start=1):
            ws.cell(r, c, h)
        hdr_row = r
        r += 1
        for row in rows:
            for c, h in enumerate(headers, start=1):
                val = row[h]
                ws.cell(r, c, val if not isinstance(val, float) else round(val, 6))
            r += 1

        for col in range(1, len(headers) + 1):
            w = 40 if col == 1 else 16
            if col >= 12:
                w = 36
            ws.column_dimensions[get_column_letter(col)].width = min(w, 52)

        ws.freeze_panes = ws.cell(hdr_row + 1, 1)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _copy_optional(primary: Path, dest: Path, written: list[Path]) -> None:
    if dest.resolve() in written:
        return
    try:
        shutil.copy2(primary, dest)
        written.append(dest.resolve())
    except OSError:
        pass


def main() -> None:
    iterations = 100
    warmup = 5
    script_dir = Path(__file__).resolve().parent

    configs: list[
        tuple[str, str, Callable[[str, int, int], tuple[dict, dict]], tuple[str, str, str]]
    ] = [
        (
            "dragoon_protocol_lv1_benchmark.xlsx",
            "Dragoon Lv-1 protocol benchmark (SIG+SIG, Setup / Attest / Verify + sizes)",
            bench_lv1_curve,
            (
                "Setup = AttKGen + VerKGen (SIG.KGen attester + SIG.KGen verifier)",
                "Attest = SIG.Sign on report message m",
                "Verify = VerSign + FinVf",
            ),
        ),
        (
            "dragoon_protocol_lv2_benchmark.xlsx",
            "Dragoon Lv-2 protocol benchmark (SIG+IBS delegated RA, Setup / Attest / Verify + sizes)",
            bench_lv2_curve,
            (
                "Setup = AttKGen + VerKGen (SIG.KGen + IBS.Setup)",
                "Attest = SIG.Sign on report message m",
                "Verify = ProxKGen + ProxSign + FinVf",
            ),
        ),
        (
            "dragoon_protocol_lv3_benchmark.xlsx",
            "Dragoon Lv-3 protocol benchmark (KBS+IBS DRAA, Setup / Attest / Verify + sizes)",
            bench_lv3_curve,
            (
                "Setup = AttKGen + VerKGen (KBS.KGen + IBS.Setup)",
                "Attest = BlGen + BlPubKey + BlSign (Fig. 3)",
                "Verify = ProxKGen + ProxSign + FinVf",
            ),
        ),
    ]

    written: list[Path] = []

    for filename, title, bench_fn, notes in configs:
        all_results: dict[str, tuple[list[dict], dict]] = {}
        for sheet, (curve, meta_extra) in _security_sheets().items():
            row, meta = bench_fn(curve, iterations=iterations, warmup=warmup)
            all_results[sheet] = ([row], {**meta, **meta_extra})

        primary = script_dir / filename
        write_protocol_workbook(primary, title, notes, all_results)
        written.append(primary.resolve())

        cwd_copy = Path.cwd() / filename
        if cwd_copy.resolve() != primary.resolve():
            _copy_optional(primary, cwd_copy, written)

        for desktop in (Path.home() / "Desktop", Path.home() / "桌面"):
            _copy_optional(primary, desktop / filename, written)

    print(f"输出目录: {script_dir}")
    print("已分别生成 Lv-1 / Lv-2 / Lv-3 表格：dragoon_protocol_lv1_benchmark.xlsx 等")
    seen_paths: set[str] = set()
    for p in written:
        key = str(p.resolve())
        if key in seen_paths:
            continue
        seen_paths.add(key)
        print(f"  {p}")


if __name__ == "__main__":
    main()
