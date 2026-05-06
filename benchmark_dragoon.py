"""
Benchmark Dragoon Lv-1 / Lv-2 / Lv-3 sub-algorithms and signature sizes; write one XLSX per level.

Each ``lambda~*_P-*`` sheet names the *approximate symmetric classical strength* (same meaning
across sheets: NIST-style mapping from ECDLP via Pollard-ρ ~ sqrt(n)).

Four supported levels (``ecdsa`` NIST curves): P-224 (~112), P-256 (~128), P-384 (~192), P-521 (~256).
There is no standard single curve in this stack for symmetric strength ~512.
"""
from __future__ import annotations

import shutil
import statistics
import time
from pathlib import Path
from typing import Callable, List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from dragoon_lv1 import DragoonLv1, DragoonLv1Evidence, run_full_protocol as run_lv1
from dragoon_lv1 import signature_byte_sizes as signature_byte_sizes_lv1
from dragoon_lv2 import DragoonLv2, DragoonLv2Evidence, run_full_protocol as run_lv2
from dragoon_lv2 import signature_byte_sizes as signature_byte_sizes_lv2
from dragoon_lv3 import (
    DragoonLv3,
    DragoonLv3Evidence,
    IBSDLMaster,
    KBSDLKeys,
    _point_bytes,
    run_full_protocol as run_lv3,
    signature_byte_sizes as signature_byte_sizes_lv3,
)


def _mean_ms(samples: List[float]) -> float:
    return statistics.mean(samples) * 1000.0


def _stdev_ms(samples: List[float]) -> float:
    if len(samples) < 2:
        return 0.0
    return statistics.stdev(samples) * 1000.0


def _bench_op(op: Callable[[], None], iterations: int) -> tuple[float, float]:
    samples: List[float] = []
    for _ in range(iterations):
        t0 = time.perf_counter()
        op()
        samples.append(time.perf_counter() - t0)
    return _mean_ms(samples), _stdev_ms(samples)


def bench_lv1_curve(curve_name: str, iterations: int, warmup: int) -> tuple[list[dict], dict]:
    dr = DragoonLv1.create(curve_name)
    m = b"tee-report-hash-placeholder-32bytes!!"

    ev, sigma_fin, _ver = run_lv1(dr, m)
    sizes = signature_byte_sizes_lv1(dr, ev, sigma_fin)

    for _ in range(warmup):
        run_lv1(dr, m)

    rows: list[dict] = []

    mean_ms, sd_ms = _bench_op(lambda: dr.att_kgen(), iterations)
    rows.append({"algorithm": "AttKGen (SIG.KGen / KBS.KGen)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    att: KBSDLKeys = dr.att_kgen()
    mean_ms, sd_ms = _bench_op(lambda: dr.kbs.sign(att.sk, m), iterations)
    rows.append({"algorithm": "SIG.Sign (attester evidence, Fig.1)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    ev_bench = dr.attest(att.sk, att.pk, m)
    mean_ms, sd_ms = _bench_op(lambda: dr.attest(att.sk, att.pk, m), iterations)
    rows.append({"algorithm": "Attest (= SIG.Sign, Fig.1)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    mean_ms, sd_ms = _bench_op(lambda: dr.att_vf(ev_bench), iterations)
    rows.append({"algorithm": "AttVf (SIG.Vf / KBS.Vf)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    mean_ms, sd_ms = _bench_op(lambda: dr.ver_kgen(), iterations)
    rows.append({"algorithm": "VerKGen (verifier SIG.KGen)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    ver = dr.ver_kgen()
    final_msg = dr._final_message(ev_bench)
    mean_ms, sd_ms = _bench_op(lambda: dr.kbs.sign(ver.sk, final_msg), iterations)
    rows.append({"algorithm": "SIG.Sign (verifier on evidence tuple, message only)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    mean_ms, sd_ms = _bench_op(lambda: dr.ver_sign(ver.sk, ev_bench), iterations)
    rows.append({"algorithm": "VerSign (AttVf + SIG.Sign, Fig.1)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    sf = dr.ver_sign(ver.sk, ev_bench)
    assert sf is not None
    mean_ms, sd_ms = _bench_op(lambda: dr.kbs.vf(ver.pk, final_msg, sf), iterations)
    rows.append({"algorithm": "SIG.Vf (verifier key, final message only)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    mean_ms, sd_ms = _bench_op(lambda: dr.fin_vf(ver.pk, ev_bench, sf), iterations)
    rows.append({"algorithm": "FinVf (AttVf + SIG.Vf, Fig.1)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    mean_ms, sd_ms = _bench_op(lambda: run_lv1(dr, m), iterations)
    rows.append(
        {
            "algorithm": "End-to-end (AttKGen+VerKGen+Attest+VerSign+FinVf)",
            "mean_ms": mean_ms,
            "stdev_ms": sd_ms,
        }
    )

    meta = {
        "curve": curve_name,
        "order_bits": dr.kbs.n.bit_length(),
        "hash": dr.kbs._hash,
        "iterations": iterations,
        **sizes,
    }
    return rows, meta


def bench_lv2_curve(curve_name: str, iterations: int, warmup: int) -> tuple[list[dict], dict]:
    dr = DragoonLv2.create(curve_name)
    m = b"tee-report-hash-placeholder-32bytes!!"

    ev, sigma_fin, _msk = run_lv2(dr, m)
    sizes = signature_byte_sizes_lv2(dr, ev, sigma_fin)

    for _ in range(warmup):
        run_lv2(dr, m)

    rows: list[dict] = []

    mean_ms, sd_ms = _bench_op(lambda: dr.att_kgen(), iterations)
    rows.append({"algorithm": "AttKGen (SIG.KGen / KBS.KGen)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    att: KBSDLKeys = dr.att_kgen()
    mean_ms, sd_ms = _bench_op(lambda: dr.kbs.sign(att.sk, m), iterations)
    rows.append({"algorithm": "SIG.Sign (attester evidence)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    mean_ms, sd_ms = _bench_op(lambda: dr.attest(att.sk, att.pk, m), iterations)
    rows.append({"algorithm": "Attest (= SIG.Sign, Fig.2 attester)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    ev_bench = dr.attest(att.sk, att.pk, m)
    mean_ms, sd_ms = _bench_op(lambda: dr.att_vf(ev_bench), iterations)
    rows.append({"algorithm": "AttVf (SIG.Vf / KBS.Vf)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    msk: IBSDLMaster = dr.ver_kgen()
    mean_ms, sd_ms = _bench_op(lambda: dr.ver_kgen(), iterations)
    rows.append({"algorithm": "VerKGen (IBS.Setup)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    mean_ms, sd_ms = _bench_op(lambda: dr.prox_kgen(msk, ev_bench.pka), iterations)
    rows.append({"algorithm": "ProxKGen (IBS.Extract)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    rk = dr.prox_kgen(msk, ev_bench.pka)
    ibs_msg = dr._ibs_message(ev_bench)
    mean_ms, sd_ms = _bench_op(lambda: dr.ibs.sign(rk, ibs_msg, msk), iterations)
    rows.append({"algorithm": "IBS.Sign (Fig.5, rk+message only)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    mean_ms, sd_ms = _bench_op(lambda: dr.prox_sign(rk, ev_bench, msk), iterations)
    rows.append({"algorithm": "ProxSign (AttVf + IBS.Sign)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    mean_ms, sd_ms = _bench_op(lambda: dr.ver_sign(msk, ev_bench), iterations)
    rows.append({"algorithm": "VerSign (AttVf + Extract + IBS.Sign)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    sf = dr.prox_sign(rk, ev_bench, msk)
    assert sf is not None
    mean_ms, sd_ms = _bench_op(lambda: dr.ibs.vf(msk, _point_bytes(ev_bench.pka), ibs_msg, sf), iterations)
    rows.append({"algorithm": "IBS.Vf (Fig.5, mpk+id+message only)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    mean_ms, sd_ms = _bench_op(lambda: dr.fin_vf(msk, ev_bench, sf), iterations)
    rows.append({"algorithm": "FinVf (AttVf + IBS.Vf)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    mean_ms, sd_ms = _bench_op(lambda: run_lv2(dr, m), iterations)
    rows.append(
        {
            "algorithm": "End-to-end (AttKGen+VerKGen+Attest+VerSign+FinVf)",
            "mean_ms": mean_ms,
            "stdev_ms": sd_ms,
        }
    )

    meta = {
        "curve": curve_name,
        "order_bits": dr.kbs.n.bit_length(),
        "hash": dr.kbs._hash,
        "iterations": iterations,
        **sizes,
    }
    return rows, meta


def bench_lv3_curve(curve_name: str, iterations: int, warmup: int) -> tuple[list[dict], dict]:
    dr = DragoonLv3.create(curve_name)
    m = b"tee-report-hash-placeholder-32bytes!!"

    ev, sigma_fin, _msk = run_lv3(dr, m)
    sizes = signature_byte_sizes_lv3(dr, ev, sigma_fin)

    for _ in range(warmup):
        run_lv3(dr, m)

    rows: list[dict] = []

    mean_ms, sd_ms = _bench_op(lambda: dr.att_kgen(), iterations)
    rows.append({"algorithm": "AttKGen (KBS.KGen)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    att: KBSDLKeys = dr.att_kgen()

    mean_ms, sd_ms = _bench_op(lambda: dr.kbs.blgen(), iterations)
    rows.append({"algorithm": "BlGen (KBS.BlGen)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    blinders = [dr.kbs.blgen() for _ in range(iterations)]
    samples = []
    for bk in blinders:
        t0 = time.perf_counter()
        dr.kbs.blpubkey(att.pk, bk)
        samples.append(time.perf_counter() - t0)
    rows.append(
        {
            "algorithm": "BlPubKey (KBS.BlPubKey, pre-sampled bk)",
            "mean_ms": _mean_ms(samples),
            "stdev_ms": _stdev_ms(samples),
        }
    )

    triples = []
    for _ in range(iterations):
        bk = dr.kbs.blgen()
        bpk = dr.kbs.blpubkey(att.pk, bk)
        triples.append((bk, bpk))
    idx = [0]

    def blsign_only():
        bk, bpk = triples[idx[0]]
        idx[0] += 1
        dr.kbs.blsign(att.sk, bk, bpk, m)

    idx[0] = 0
    mean_ms, sd_ms = _bench_op(blsign_only, iterations)
    rows.append({"algorithm": "BlSign (KBS.BlSign, precomputed bk,bpk)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    mean_ms, sd_ms = _bench_op(lambda: dr.attest(att.sk, att.pk, m), iterations)
    rows.append({"algorithm": "Attest (BlGen+BlPubKey+BlSign, Fig.3)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    bk_fix = triples[0][0]
    bpk_fix = triples[0][1]
    sigma_a = dr.kbs.blsign(att.sk, bk_fix, bpk_fix, m)
    ev_bench = DragoonLv3Evidence(bpk=bpk_fix, m=m, sigma_a=sigma_a)

    mean_ms, sd_ms = _bench_op(lambda: dr.att_vf(ev_bench.bpk, ev_bench.m, ev_bench.sigma_a), iterations)
    rows.append({"algorithm": "AttVf (KBS.Vf)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    msk: IBSDLMaster = dr.ver_kgen()
    mean_ms, sd_ms = _bench_op(lambda: dr.ver_kgen(), iterations)
    rows.append({"algorithm": "VerKGen (IBS.Setup)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    mean_ms, sd_ms = _bench_op(lambda: dr.prox_kgen(msk, ev_bench.bpk), iterations)
    rows.append({"algorithm": "ProxKGen (IBS.Extract)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    rk = dr.prox_kgen(msk, ev_bench.bpk)
    ibs_msg = dr._ibs_message(ev_bench)
    mean_ms, sd_ms = _bench_op(lambda: dr.ibs.sign(rk, ibs_msg, msk), iterations)
    rows.append({"algorithm": "IBS.Sign (Fig.5, rk+message only)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    mean_ms, sd_ms = _bench_op(lambda: dr.prox_sign(rk, ev_bench, msk), iterations)
    rows.append({"algorithm": "ProxSign (AttVf + IBS.Sign)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    mean_ms, sd_ms = _bench_op(lambda: dr.ver_sign(msk, ev_bench), iterations)
    rows.append({"algorithm": "VerSign (AttVf + Extract + IBS.Sign)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    sf = dr.prox_sign(rk, ev_bench, msk)
    assert sf is not None
    mean_ms, sd_ms = _bench_op(lambda: dr.ibs.vf(msk, _point_bytes(ev_bench.bpk), ibs_msg, sf), iterations)
    rows.append({"algorithm": "IBS.Vf (Fig.5, mpk+id+message only)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    mean_ms, sd_ms = _bench_op(lambda: dr.fin_vf(msk, ev_bench, sf), iterations)
    rows.append({"algorithm": "FinVf (AttVf + IBS.Vf)", "mean_ms": mean_ms, "stdev_ms": sd_ms})

    mean_ms, sd_ms = _bench_op(lambda: run_lv3(dr, m), iterations)
    rows.append(
        {
            "algorithm": "End-to-end (AttKGen+VerKGen+Attest+VerSign+FinVf)",
            "mean_ms": mean_ms,
            "stdev_ms": sd_ms,
        }
    )

    meta = {
        "curve": curve_name,
        "order_bits": dr.kbs.n.bit_length(),
        "hash": dr.kbs._hash,
        "iterations": iterations,
        **sizes,
    }
    return rows, meta


def write_workbook(path: Path, title_line: str, all_results: dict[str, tuple[list[dict], dict]]) -> None:
    wb = Workbook()
    first = True
    for sheet_name, (rows, meta) in all_results.items():
        if first:
            ws = wb.active
            ws.title = sheet_name[:31]
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name[:31])

        ws["A1"] = title_line
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        r = 3
        ws.cell(r, 1, "Parameter")
        ws.cell(r, 2, "Value")
        r += 1
        for k, v in meta.items():
            ws.cell(r, 1, str(k))
            ws.cell(r, 2, v if not isinstance(v, float) else round(v, 6))
            r += 1

        r += 1
        ws.cell(r, 1, "Sub-algorithm")
        ws.cell(r, 2, "mean time (ms)")
        ws.cell(r, 3, "stdev (ms)")
        r += 1
        for row in rows:
            ws.cell(r, 1, row["algorithm"])
            ws.cell(r, 2, round(row["mean_ms"], 6))
            ws.cell(r, 3, round(row["stdev_ms"], 6))
            r += 1

        for col in range(1, 4):
            ws.column_dimensions[get_column_letter(col)].width = 52 if col == 1 else 18

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _security_sheets() -> dict[str, tuple[str, dict]]:
    return {
        "lambda~112_P-224": ("SECP224R1", {"symmetric_security_bits_approx": 112}),
        "lambda~128_P-256": ("SECP256R1", {"symmetric_security_bits_approx": 128}),
        "lambda~192_P-384": ("SECP384R1", {"symmetric_security_bits_approx": 192}),
        "lambda~256_P-521": ("SECP521R1", {"symmetric_security_bits_approx": 256}),
    }


def _build_results(bench_fn: Callable[[str, int, int], tuple[list[dict], dict]], iterations: int, warmup: int) -> dict:
    all_results: dict[str, tuple[list[dict], dict]] = {}
    for sheet, (curve, meta_extra) in _security_sheets().items():
        rows, meta = bench_fn(curve, iterations=iterations, warmup=warmup)
        meta = {**meta, **meta_extra}
        all_results[sheet] = (rows, meta)
    return all_results


def _emit_workbook(
    script_dir: Path,
    filename: str,
    title_line: str,
    all_results: dict[str, tuple[list[dict], dict]],
    written: list[Path],
) -> None:
    primary = script_dir / filename
    write_workbook(primary, title_line, all_results)
    written.append(primary.resolve())

    cwd_copy = Path.cwd() / filename
    if cwd_copy.resolve() != primary.resolve():
        shutil.copy2(primary, cwd_copy)
        written.append(cwd_copy.resolve())

    for desktop in (Path.home() / "Desktop", Path.home() / "桌面"):
        dest = desktop / filename
        if desktop.is_dir() and dest.resolve() not in written:
            shutil.copy2(primary, dest)
            written.append(dest.resolve())


def main() -> None:
    iterations = 100
    warmup = 5
    written: list[Path] = []
    script_dir = Path(__file__).resolve().parent

    _emit_workbook(
        script_dir,
        "dragoon_lv1_benchmark.xlsx",
        "Lv-1 Dragoon benchmark (SIG + SIG, vanilla RA, Fig.1)",
        _build_results(bench_lv1_curve, iterations, warmup),
        written,
    )
    _emit_workbook(
        script_dir,
        "dragoon_lv2_benchmark.xlsx",
        "Lv-2 Dragoon benchmark (SIG + IBS, delegated RA, Fig.2)",
        _build_results(bench_lv2_curve, iterations, warmup),
        written,
    )
    _emit_workbook(
        script_dir,
        "dragoon_lv3_benchmark.xlsx",
        "Lv-3 Dragoon benchmark (KBSDL + IBSDL, Fig.3)",
        _build_results(bench_lv3_curve, iterations, warmup),
        written,
    )

    print("已生成表格，绝对路径如下（任一路径均可打开）:")
    seen: set[str] = set()
    for p in written:
        key = str(p.resolve())
        if key in seen:
            continue
        seen.add(key)
        print(f"  {p}")


if __name__ == "__main__":
    main()
